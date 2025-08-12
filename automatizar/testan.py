import os
import re
import pytesseract
from pdf2image import convert_from_path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# Configuração Tesseract
tesseract_config = '--psm 6'
tesseract_lang = 'por'

# Palavras-chave para detecção
DEFINIR_NF = {"Prefeitura", "Nota Fiscal", "Nota de Serviço", "Recibo"}
DEFINIR_BOLETO = {"Linha Digitável", "Código de Barras", "Agência/Código do Beneficiário", "Linha Digitavel", "Agência", "Código do Beneficiário"}

# Campos por tipo
CAMPOS_BOLETO = {
    "Nº do Documento": r"\d+",
    "Vencimento": r"\d{2}/\d{2}/\d{4}",
    "Valor do Documento": r"\d{1,3}(?:\.\d{3})*,\d{2}"
}

CAMPOS_NF = {
    "Número da Nota": r"\d+",
    "Data de Emissão": r"\d{2}/\d{2}/\d{4}(?:\s+\d{2}:\d{2}(?::\d{2})?)?",
    "Data e Hora de Emissão": r"\d{2}/\d{2}/\d{4}(?:\s+\d{2}:\d{2}(?::\d{2})?)?",
    "Valor Total da Nota": r"R?\$?\s*\d{1,3}(?:\.\d{3})*,\d{2}"
}

# Detectar tipo do documento
def detectar_tipo_documento(texto_continuo):
    texto_lower = texto_continuo.lower()
    if any(p.lower() in texto_lower for p in DEFINIR_BOLETO):
        return "BOLETO"
    if any(p.lower() in texto_lower for p in DEFINIR_NF):
        return "NF"
    return "BOLETO"

# Renomear arquivo
def renomear_pdf(caminho, tipo):
    pasta, nome_arquivo = os.path.split(caminho)
    nome, ext = os.path.splitext(nome_arquivo)
    if not nome.endswith(f"_{tipo}"):
        novo_nome = f"{nome}_{tipo}{ext}"
        novo_caminho = os.path.join(pasta, novo_nome)
        os.rename(caminho, novo_caminho)
        print(f"📂 Arquivo renomeado para: {novo_nome}")
        return novo_caminho
    return caminho

# Extrair campos
def extrair_campos(texto_continuo, campos):
    encontrados = {}
    for chave, padrao_base in campos.items():
        if re.search(re.escape(chave), texto_continuo, re.IGNORECASE):
            padrao = re.search(
                rf"{re.escape(chave)}(?:\s+\S+){{0,5}}?\s+({padrao_base})",
                texto_continuo,
                re.IGNORECASE
            )
            encontrados[chave] = padrao.group(1).strip() if padrao else "Não encontrado"
    return encontrados

# Carregar registros já existentes no Excel
def carregar_registros_existentes(arquivo_excel):
    registros = set()
    if os.path.exists(arquivo_excel):
        try:
            wb = load_workbook(arquivo_excel)
            ws = wb.active
            headers = {cell.value: idx for idx, cell in enumerate(ws[1])}
            for row in ws.iter_rows(min_row=2, values_only=True):
                num_doc = None
                data_doc = None
                for col_name in ["Nº do Documento", "Número da Nota"]:
                    if col_name in headers:
                        num_doc = row[headers[col_name]]
                        if num_doc:
                            break
                for col_name in ["Data de Emissão", "Data e Hora de Emissão", "Vencimento"]:
                    if col_name in headers:
                        data_doc = row[headers[col_name]]
                        if data_doc:
                            break
                if num_doc and data_doc:
                    registros.add((str(num_doc).strip(), str(data_doc).strip()))
        except Exception as e:
            print(f"⚠️ Erro ao ler Excel existente: {e}")
    return registros

# Processar PDF individual
def processar_pdf(pdf_path):
    print(f"\n🔍 Processando: {os.path.basename(pdf_path)}...")
    resultado = {
        "Arquivo Original": os.path.basename(pdf_path),
        "Arquivo Renomeado": "",
        "Tipo": "",
        "Campos": {}
    }
    try:
        imagens = convert_from_path(pdf_path)
        tipo_documento = None
        campos_referencia = {}
        for i, imagem in enumerate(imagens[:2]):  # só primeiras 2 páginas
            texto = pytesseract.image_to_string(imagem, config=tesseract_config, lang=tesseract_lang)
            texto_continuo = " ".join(texto.split())
            if tipo_documento is None:
                tipo_documento = detectar_tipo_documento(texto_continuo)
                resultado["Tipo"] = tipo_documento
                campos_referencia = CAMPOS_NF if tipo_documento == "NF" else CAMPOS_BOLETO
                novo_caminho = renomear_pdf(pdf_path, tipo_documento)
                resultado["Arquivo Renomeado"] = os.path.basename(novo_caminho)
            if i == 0:
                campos_encontrados = extrair_campos(texto_continuo, campos_referencia)
                resultado["Campos"].update(campos_encontrados)
        return resultado
    except Exception as e:
        print(f"❌ Erro ao processar {pdf_path}: {e}")
        resultado["Tipo"] = f"ERRO: {e}"
        return resultado

# Exportar para Excel
def exportar_para_excel(resultados, pasta_saida, nome_arquivo):
    excel_path = os.path.join(pasta_saida, nome_arquivo)
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
        cabecalhos = [cell.value for cell in ws[1]]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados PDF"
        cabecalhos = ["Arquivo Original", "Arquivo Renomeado", "Tipo"]
        campos_unicos = set()
        for res in resultados:
            campos_unicos.update(res["Campos"].keys())
        cabecalhos.extend(sorted(campos_unicos))
        ws.append(cabecalhos)
        for cell in ws[1]:
            cell.font = Font(bold=True)
    for res in resultados:
        linha = [
            res["Arquivo Original"],
            res.get("Arquivo Renomeado", ""),
            res["Tipo"]
        ]
        for campo in cabecalhos[3:]:
            linha.append(res["Campos"].get(campo, "Não encontrado"))
        ws.append(linha)
    wb.save(excel_path)
    print(f"📊 Resultados exportados para: {excel_path}")

# Processar pasta inteira
def processar_pasta(pasta_entrada, pasta_saida):
    if not os.path.exists(pasta_saida):
        os.makedirs(pasta_saida)
    excel_controle = os.path.join(pasta_saida, "resultados_pdfs.xlsx")
    registros_existentes = carregar_registros_existentes(excel_controle)
    pdfs = [f for f in os.listdir(pasta_entrada) if f.lower().endswith(".pdf")]
    resultados = []
    for pdf in pdfs:
        pdf_path = os.path.join(pasta_entrada, pdf)
        res = processar_pdf(pdf_path)
        num_doc = res["Campos"].get("Nº do Documento") or res["Campos"].get("Número da Nota")
        data_doc = res["Campos"].get("Data de Emissão") or res["Campos"].get("Data e Hora de Emissão") or res["Campos"].get("Vencimento")
        if num_doc and data_doc and (str(num_doc).strip(), str(data_doc).strip()) in registros_existentes:
            print(f"⏭️ Pulando {pdf} — já processado anteriormente.")
            continue
        resultados.append(res)
    if resultados:
        exportar_para_excel(resultados, pasta_saida, "resultados_pdfs.xlsx")
    else:
        print("⚠️ Nenhum PDF novo para exportar.")

if __name__ == "__main__":
    PASTA_PDFS = "automatizar/BOLETOS"
    PASTA_SAIDA = "automatizar/Exel"
    try:
        pytesseract.get_tesseract_version()
    except:
        print("❌ Tesseract OCR não está instalado ou não está no PATH.")
        exit()
    processar_pasta(PASTA_PDFS, PASTA_SAIDA)
