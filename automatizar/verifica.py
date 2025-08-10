#!/usr/bin/env python3
import os
import datetime
import openpyxl as px

Notas = "automatizar/test.xlsx"

def inicializar_arquivo():
    if not os.path.exists(Notas):
        print(f"Arquivo '{Notas}' ainda não criado.")
        exit(1)

def buscar_mes_e_ano_anterior():
    hoje = datetime.date.today()
    primeiro_dia_mes_atual = hoje.replace(day=1)
    ultimo_dia_mes_passado = primeiro_dia_mes_atual - datetime.timedelta(days=1)
    return ultimo_dia_mes_passado.month, hoje.year  # mês anterior, ano atual

def ler():
    print("\n📄 Lista de registros do MÊS ANTERIOR (do ano atual):\n")
    Base = px.load_workbook(Notas, data_only=True)
    planilha = Base.active

    headers = [cell.value for cell in planilha[1]]
    apelido_idx = headers.index('Apelido')
    codigo_idx = headers.index('chave')
    descricao_idx = headers.index('Descritivo Pagamento:')
    mes_idx = headers.index('Mês')
    ano_idx = headers.index('Ano')

    mes_anterior, ano_atual = buscar_mes_e_ano_anterior()

    for i, row in enumerate(planilha.iter_rows(min_row=2, values_only=True), start=2):
        apelido = row[apelido_idx]
        codigo = row[codigo_idx]
        descricao = row[descricao_idx]
        campo_mes = row[mes_idx]
        campo_ano = row[ano_idx]

        if campo_mes == mes_anterior and campo_ano == ano_atual:
            if apelido and codigo and descricao:
                print(f"{i} | {apelido} | {codigo} | {descricao}")
            else:
                print(f"{i} | Dados incompletos na linha {i}")

def verificar_pdfs_existentes(pasta_pdfs="."):
    apelidos_xlsx = set()

    # Dados da data
    mes_anterior, ano_atual = buscar_mes_e_ano_anterior()

    Base = px.load_workbook(Notas, data_only=True)
    planilha = Base.active

    headers = [cell.value for cell in planilha[1]]
    apelido_idx = headers.index('Apelido')
    mes_idx = headers.index('Mês')
    ano_idx = headers.index('Ano')

    for row in planilha.iter_rows(min_row=2, values_only=True):
        apelido = row[apelido_idx]
        campo_mes = row[mes_idx]
        campo_ano = row[ano_idx]

        if campo_mes == mes_anterior and campo_ano == ano_atual:
            if apelido:
                apelidos_xlsx.add(str(apelido).strip().lower())

    if not apelidos_xlsx:
        print("⚠️ Nenhum apelido encontrado para o mês anterior na planilha.")
        return False

    # Lista arquivos PDF na pasta
    arquivos_pdf = [
        f for f in os.listdir(pasta_pdfs)
        if f.lower().endswith(".pdf")
    ]
    nomes_pdfs = set(
        os.path.splitext(f)[0].strip().lower()
        for f in arquivos_pdf
    )

    faltando = apelidos_xlsx - nomes_pdfs

    if faltando:
        print("\n❌ Os seguintes apelidos do mês anterior NÃO possuem PDF correspondente:")
        for apelido in sorted(faltando):
            print(f"- {apelido}.pdf")
        print("\n⚠️ Certifique-se de que os PDFs estejam na pasta correta.")
        return False
    else:
        print("✅ Todos os PDFs dos apelidos do mês anterior estão presentes.")
        return True

def main():
    print("📁 Verificador de PDFs por Apelido")
    print("===================================")
    print("\nDeseja continuar? (Y/n) > ", end="")
    resposta = input().strip().lower()

    if resposta == 'y' or resposta == '':
        inicializar_arquivo()
        ler()
        verificar_pdfs_existentes()
        print("\n✅ Operação concluída.\n")
    else:
        print("🚫 Operação cancelada. Saindo do programa.")

if __name__ == "__main__":
    main()
