import os
import pytesseract
import re
import json
import pandas as pd
from pdf2image import convert_from_path
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime, timedelta

# --- ConfiguraÃ§Ãµes globais ---
tesseract_config = '--psm 6'
tesseract_lang = 'por'
# pytesseract.pytesseract.tesseract_cmd = r"C:\Users\abrito\AppData\Local\Programs\Tesseract-OCR\tesseract.exe"
# poppler_path = r"C:\Users\abrito\Documents\Release-24.08.0-0\poppler-24.08.0\Library\bin"

# DefiniÃ§Ãµes de documentos
DEFINIR_NF = {"Prefeitura", "Nota Fiscal", "Nota de ServiÃ§o", "Recibo"}
DEFINIR_BOLETO = {"Linha DigitÃ¡vel", "CÃ³digo de Barras", "AgÃªncia/CÃ³digo do BeneficiÃ¡rio", "Linha Digitavel", "AgÃªncia", "CÃ³digo do BeneficiÃ¡rio"}

CAMPOS_BOLETO = {
    "NÃºmero do Documento": r"\d+",
    "Vencimento": r"\d{2}/\d{2}/\d{4}",
    "Valor do Documento": r"\d{1,3}(?:\.\d{3})*,\d{2}"
}

CAMPOS_NF = {
    "NÃºmero da Nota": r"\d+",
    "Data de EmissÃ£o": r"\d{2}/\d{2}/\d{4}(?:\s+\d{2}:\d{2}(?::\d{2})?)?",
    "Data e Hora de EmissÃ£o": r"\d{2}/\d{2}/\d{4}(?:\s+\d{2}:\d{2}(?::\d{2})?)?",
    "Valor Total da Nota": r"R?\$?\s*\d{1,3}(?:\.\d{3})*,\d{2}",
    "Valor Total do ServiÃ§o": r"R?\$?\s*\d{1,3}(?:\.\d{3})*,\d{2}"
}

# --- FunÃ§Ãµes do histÃ³rico ---

def limpar_log_antigo(caminho_log, dias_retencao=45):
    """
    Remove registros do log com mais de X dias (padrÃ£o: 45 dias)
    Retorna True se houve alteraÃ§Ã£o no arquivo, False caso contrÃ¡rio
    """
    if not os.path.exists(caminho_log):
        return False
    
    try:
        # Data limite para manutenÃ§Ã£o (hoje - dias_retencao)
        data_limite = datetime.now() - timedelta(days=dias_retencao)
        registros_atualizados = []
        alterado = False
        
        with open(caminho_log, 'r', encoding='utf-8') as f:
            for line in f:
                try:
                    registro = json.loads(line.strip())
                    # Verifica se o registro tem data_processo e se Ã© antigo
                    if 'data_processo' in registro:
                        data_registro = datetime.strptime(registro['data_processo'], "%Y-%m-%d %H:%M:%S")
                        if data_registro >= data_limite:
                            registros_atualizados.append(registro)
                        else:
                            alterado = True
                    else:
                        # MantÃ©m registros sem data (para backward compatibility)
                        registros_atualizados.append(registro)
                except (json.JSONDecodeError, ValueError) as e:
                    print(f"âš ï¸ Erro ao processar linha do log: {e}")
                    continue
        
        # Se houve registros removidos, reescreve o arquivo
        if alterado:
            with open(caminho_log, 'w', encoding='utf-8') as f:
                for registro in registros_atualizados:
                    f.write(json.dumps(registro, ensure_ascii=False) + "\n")
            print(f"â™»ï¸ Log limpo: removidos registros com mais de {dias_retencao} dias")
        
        return alterado
    
    except Exception as e:
        print(f"âŒ Erro ao limpar log: {str(e)}")
        return False

def ler_log(caminho_log):
    historico = []
    if os.path.exists(caminho_log):
        try:
            with open(caminho_log, 'r', encoding='utf-8') as f:
                for line in f:
                    historico.append(json.loads(line.strip()))
        except (IOError, json.JSONDecodeError) as e:
            print(f"âš ï¸ Erro ao ler ou decodificar o arquivo de histÃ³rico: {e}")
            historico = []
    return historico

def adicionar_ao_historico(caminho_log, resultado):
    with open(caminho_log, 'a', encoding='utf-8') as f:
        log_data = {
            'arquivo': resultado['Arquivo Original'],
            'tipo': resultado['Tipo'],
            'campos_extraidos': resultado['Campos'],
            'data_processo': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        f.write(json.dumps(log_data, ensure_ascii=False) + "\n")

# --- FunÃ§Ãµes de verificaÃ§Ã£o ---

def documento_ja_registrado(dados_log, tipo, campos):
    """Verifica se documento com mesmo tipo e campos-chave jÃ¡ existe no log."""
    for log_item in dados_log:
        if log_item.get('tipo') != tipo:
            continue
        if tipo == 'NF':
            if (log_item['campos_extraidos'].get('Data e Hora de EmissÃ£o') == campos.get('Data e Hora de EmissÃ£o') and
                log_item['campos_extraidos'].get('NÃºmero da Nota') == campos.get('NÃºmero da Nota')):
                return True
        elif tipo == 'BOLETO':
            if (log_item['campos_extraidos'].get('Vencimento') == campos.get('Vencimento') and
                log_item['campos_extraidos'].get('NÃºmero do Documento') == campos.get('NÃºmero do Documento')):
                return True
    return False

# --- FunÃ§Ãµes de processamento ---

def renomear_pdf(caminho, tipo):
    pasta, nome_arquivo = os.path.split(caminho)
    nome, ext = os.path.splitext(nome_arquivo)
    if not nome.endswith(f"_{tipo}"):
        novo_nome = f"{nome}_{tipo}{ext}"
        novo_caminho = os.path.join(pasta, novo_nome)
        os.rename(caminho, novo_caminho)
        print(f"ðŸ“‚ Arquivo renomeado para: {novo_nome}")
        return novo_caminho
    return caminho

def detectar_tipo_documento(texto_continuo):
    texto_lower = texto_continuo.lower()
    if any(p.lower() in texto_lower for p in DEFINIR_BOLETO):
        return "BOLETO"
    if any(p.lower() in texto_lower for p in DEFINIR_NF):
        return "NF"
    return "BOLETO"

def extrair_campos(texto_continuo, campos):
    encontrados = {}
    for chave, padrao_base in campos.items():
        if re.search(re.escape(chave), texto_continuo, re.IGNORECASE):
            padrao = re.search(
                rf"{re.escape(chave)}.*?({padrao_base})", 
                texto_continuo, 
                re.IGNORECASE | re.DOTALL
            )
            encontrados[chave] = padrao.group(1).strip() if padrao else "NÃ£o encontrado"
    return encontrados

def processar_pdf(pdf_path):
    print(f"\nðŸ” Processando o arquivo: {os.path.basename(pdf_path)}...")
    resultado = {
        'Arquivo Original': os.path.basename(pdf_path),
        'Arquivo Renomeado': '',
        'Tipo': '',
        'Campos': {}
    }
    try:
        imagens = convert_from_path(pdf_path)  # poppler_path=poppler_path
        tipo_documento = None
        campos_referencia = {}
        
        for i, imagem in enumerate(imagens[:2]):
            texto = pytesseract.image_to_string(imagem, config=tesseract_config, lang=tesseract_lang)
            texto_continuo = " ".join(texto.split())
            
            if tipo_documento is None:
                tipo_documento = detectar_tipo_documento(texto_continuo)
                resultado['Tipo'] = tipo_documento
                campos_referencia = CAMPOS_NF if tipo_documento == "NF" else CAMPOS_BOLETO
                
                novo_caminho = renomear_pdf(pdf_path, tipo_documento)
                resultado['Arquivo Renomeado'] = os.path.basename(novo_caminho)
            
            if i == 0:
                campos_encontrados = extrair_campos(texto_continuo, campos_referencia)
                resultado['Campos'].update(campos_encontrados)
        
        print(f"âœ… Processado: {resultado['Tipo']}")
        print(f"   Campos encontrados: {', '.join([f'{k}: {v}' for k, v in resultado['Campos'].items()])}")
        return resultado
        
    except Exception as e:
        print(f"âŒ Erro ao processar {pdf_path}: {str(e)}")
        resultado['Tipo'] = f"ERRO: {str(e)}"
        return resultado

def exportar_para_excel(resultados, pasta_saida):
    if not resultados:
        print("âš ï¸ Nenhum resultado para exportar!")
        return None
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados PDF"
    
    cabecalhos = ['Arquivo Original', 'Arquivo Renomeado', 'Tipo']
    campos_unicos = set()
    
    for res in resultados:
        campos_unicos.update(res['Campos'].keys())
    
    cabecalhos.extend(sorted(campos_unicos))
    ws.append(cabecalhos)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    for res in resultados:
        linha = [
            res['Arquivo Original'],
            res.get('Arquivo Renomeado', ''),
            res['Tipo']
        ]
        for campo in cabecalhos[3:]:
            linha.append(res['Campos'].get(campo, "NÃ£o encontrado"))
        ws.append(linha)
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(pasta_saida, f"resultados_pdfs_{timestamp}.xlsx")
    wb.save(excel_path)
    print(f"\nðŸ“Š Resultados exportados para: {excel_path}")
    return excel_path

def processar_pasta(pasta_entrada, pasta_saida=None):
    if pasta_saida is None:
        pasta_saida = pasta_entrada
    
    if not os.path.exists(pasta_entrada):
        print(f"âŒ Pasta de entrada nÃ£o encontrada: {pasta_entrada}")
        return
    
    if not os.path.exists(pasta_saida):
        os.makedirs(pasta_saida)

    caminho_historico = os.path.join(pasta_saida, "arquivos_processados.log")
    limpar_log_antigo(caminho_historico)
    historico_processados = ler_log(caminho_historico)

    pdfs = [f for f in os.listdir(pasta_entrada) if f.lower().endswith('.pdf')]
    if not pdfs:
        print(f"âŒ Nenhum arquivo PDF encontrado em: {pasta_entrada}")
        return
    
    print(f"\nðŸ“‚ Encontrados {len(pdfs)} arquivos PDF para processar...")
    
    resultados = []
    for pdf in pdfs:
        pdf_path = os.path.join(pasta_entrada, pdf)

        # Processa sempre para extrair dados
        resultado = processar_pdf(pdf_path)

        # Se jÃ¡ existe no log com mesmos campos, pula exportaÃ§Ã£o
        if documento_ja_registrado(historico_processados, resultado['Tipo'], resultado['Campos']):
            print(f"â© Documento jÃ¡ registrado no log, pulando exportaÃ§Ã£o: {pdf}")
            continue

        if not resultado['Tipo'].startswith("ERRO") and resultado['Campos']:
            adicionar_ao_historico(caminho_historico, resultado)
            resultados.append(resultado)
            
    exportar_para_excel(resultados, pasta_saida)

if __name__ == "__main__":
    PASTA_PDFS = "automacao-main/Boletos"
    PASTA_SAIDA = "automacao-main/Excel"
    
    try:
        pytesseract.get_tesseract_version()
    except:
        print("âŒ Tesseract OCR nÃ£o estÃ¡ instalado ou nÃ£o estÃ¡ no PATH")
        exit()
    
    processar_pasta(PASTA_PDFS, PASTA_SAIDA)


#adicinar pensamento depois

from openpyxl import Workbook, load_workbook

def exportar_para_excel(resultados, pasta_saida, nome_arquivo="resultados_pdfs.xlsx"):
    if not resultados:
        print("âš ï¸ Nenhum resultado para exportar!")
        return None

    excel_path = os.path.join(pasta_saida, nome_arquivo)

    # Se jÃ¡ existir, abre o arquivo existente
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        if "Resultados PDF" in wb.sheetnames:
            ws = wb["Resultados PDF"]
        else:
            ws = wb.active
    else:
        # Cria novo workbook e adiciona cabeÃ§alhos
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados PDF"
        cabecalhos = ['Arquivo Original', 'Arquivo Renomeado', 'Tipo']
        campos_unicos = set()
        for res in resultados:
            campos_unicos.update(res['Campos'].keys())
        cabecalhos.extend(sorted(campos_unicos))
        ws.append(cabecalhos)
        for cell in ws[1]:
            cell.font = Font(bold=True)

    # Descobre campos existentes no cabeÃ§alho
    cabecalhos_existentes = [cell.value for cell in ws[1]]

    for res in resultados:
        linha = [
            res['Arquivo Original'],
            res.get('Arquivo Renomeado', ''),
            res['Tipo']
        ]
        for campo in cabecalhos_existentes[3:]:
            linha.append(res['Campos'].get(campo, "NÃ£o encontrado"))
        ws.append(linha)

    # Ajusta largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(excel_path)
    print(f"\nðŸ“Š Resultados adicionados em: {excel_path}")
    return excel_path
